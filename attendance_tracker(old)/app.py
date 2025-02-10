from flask import Flask, render_template, request, redirect, url_for, session
import json
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.secret_key = 'basta'

# Load student list from JSON file
def load_student_list():
    with open('data/students.json', 'r') as file:
        return json.load(file)

# Save student list to JSON file
def save_student_list(student_list):
    with open('data/students.json', 'w') as file:
        json.dump(student_list, file, indent=4)

# Load admin credentials from JSON file
def load_admin_credentials():
    with open('data/credentials.json', 'r') as file:
        return json.load(file)

STUDENT_LIST = load_student_list()
ADMIN_CREDENTIALS = load_admin_credentials()

@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        password = request.form['password']

        if password in ADMIN_CREDENTIALS:
            session['section'] = ADMIN_CREDENTIALS[password]
            return redirect(url_for('section_dashboard'))
        else:
            return render_template('login.html', error="Invalid password")

    return render_template('login.html')

@app.route('/section', methods=['GET', 'POST'])
def section_dashboard():
    if 'section' not in session:
        return redirect(url_for('login'))

    section = session['section']
    student_list = STUDENT_LIST[section]

    if request.method == 'POST':
        attendance_data = []
        for gender in ['male', 'female']:
            for student in student_list[gender]:
                status = request.form.get(f'status_{student}', 'absent')
                if status == 'late':
                    time_late = request.form.get(f'time_{student}', 'N/A')
                else:
                    time_late = 'N/A'
                attendance_data.append({
                    'student': student,
                    'status': status,
                    'time_late': time_late
                })

        # Save attendance to Excel
        save_attendance_to_excel(attendance_data, section)
        return redirect(url_for('section_dashboard'))

    return render_template('section_dashboard.html', section=section, student_list=student_list)

@app.route('/admin', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        if username in ADMIN_CREDENTIALS and ADMIN_CREDENTIALS[username] == password:
            session['admin'] = True
            return redirect(url_for('admin_dashboard'))
    
    return render_template('admin_login.html')

@app.route('/admin/dashboard')
def admin_dashboard():
    if not session.get('admin'):
        return redirect(url_for('admin_login'))
    
    # Load attendance records and students data
    attendance_records = load_attendance_records()
    return render_template('admin_dashboard.html', attendance_records=attendance_records, student_list=STUDENT_LIST)

# Helper function to load attendance records
def load_attendance_records():
    excel_filename = 'attendance_records.xlsx'
    
    # Check if the file exists and load it, otherwise return an empty list
    if os.path.exists(excel_filename):
        wb = load_workbook(excel_filename)
        ws = wb["Attendance"]
        
        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            records.append({
                'date': row[0],
                'time': row[1],
                'student': row[2],
                'status': row[3],
                'time_late': row[4]
            })
        return records
    
    return []

# Route to add/change students
@app.route('/admin/students', methods=['POST'])
def admin_modify_students():
    if not session.get('admin'):
        return redirect(url_for('admin_login'))
    
    if request.method == 'POST':
        new_student = request.form.get('new_student')
        gender = request.form.get('gender')
        section = request.form.get('section')
        if new_student and gender in ['male', 'female'] and section in STUDENT_LIST:
            STUDENT_LIST[section][gender].append(new_student)
            save_student_list(STUDENT_LIST)
    
    return redirect(url_for('admin_dashboard'))

def save_attendance_to_excel(attendance_data, section):
    excel_filename = f'attendance_records_{section}.xlsx'
    
    # Check if the file exists and load it, otherwise create a new one
    if os.path.exists(excel_filename):
        wb = load_workbook(excel_filename)
    else:
        wb = Workbook()
        wb.remove(wb.active)  # Remove the default sheet
        wb.create_sheet(title="Attendance")
        ws = wb["Attendance"]
        ws.append(['Date', 'Time', 'Student', 'Status', 'Time Late'])

    ws = wb["Attendance"]

    for data in attendance_data:
        ws.append([
            datetime.now().strftime('%Y-%m-%d'),
            datetime.now().strftime('%H:%M:%S'),
            data['student'],
            data['status'],
            data['time_late']
        ])

    wb.save(excel_filename)
    print(f"Attendance data saved to '{excel_filename}' successfully.")

if __name__ == '__main__':
    app.run(debug=True)

